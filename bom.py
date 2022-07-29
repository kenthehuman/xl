"""
App to create a new excel sheet Bill of Materials
User inputs the components needed and python exports complete .xlsx file

Idea to add, copy from existing BOM or template
"""
from functools import partial
from openpyxl import Workbook


# new blank Workbook object
wb = Workbook()
# Get active sheet from workbook
sheet = wb.active

# Note: The first row or column integer
# is 1, not 0. Cell object is created by
# using sheet object's cell() method.
c1 = sheet.cell(row = 1, column = 1)

# writing values to cellse
c1.value = "Hello"

c2 = sheet.cell(row= 1 , column = 2) # Set new cell location
c2.value = "World"

# Once have a Worksheet object, one can
# access a cell object by its name also.
# A2 means column = 1 & row = 2.
c3 = sheet['A2']
c3.value = "Welcome"

# B2 means column = 2 & row = 2.
c4 = sheet['B2']
c4.value = "Everyone"

# Anytime you modify the Workbook object
# or its sheets and cells, the spreadsheet
# file will not be saved until you call
# the save() workbook method.
wb.save("sample.xlsx")

# Variables for the Template ie. Part Number, Description, etc
part_number, part_description = '', ''
part_title = ['Part Number', part_number, part_description, 'Bill of Materials']
table_header = ['#', 'Component', 'Description', 'Material', 'Qty', 'Part Number', 'Supplier']

for i in range(len(table_header)):
    cell = sheet.cell(row=1, column= i+1)
    cell.value = table_header[i]

wb.save("sample.xlsx")

# >>> for i in range(len(a)):
# ...     print(a[i])