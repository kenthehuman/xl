

from openpyxl import Workbook
from openpyxl import load_workbook

# workbook = Workbook()
# sheet = workbook.active
"""
sheet["A1"] = "hello"
sheet["B1"] = "world!"

workbook.save(filename="hello_world.xlsx")
"""



workbook = load_workbook(filename="hello_world.xlsx")
workbook.sheetnames


sheet = workbook.active
# sheet
"""

sheet.title
"""

# sheet["A1"]


# print(sheet["A1"].value)


# sheet["F10"].value

# sheet.cell(row=10, column=6)


# sheet.cell(row=10, column=6).value

# print(sheet["A1:C2"]) #slice colums

# Get all cells from column A
# print(sheet["A"])

# Get all cells for a range of columns
# print(sheet["A:B"])

# Get all cells from row 5
# print(sheet[5])

# for row in sheet.iter_rows(min_row=1, \
#     max_row=5,\
#         min_col=1, \
#             max_col=4):
#             print(row)

print("rows")
for row in sheet.rows:
    print(row)
    row.value = row

print('\ncolumns')
for colum in sheet.columns:
    print(colum)
    colum.value = colum

workbook.save(filename="hello_world.xlsx")

# Python program to read an excel file

# import openpyxl module
import openpyxl

# Give the location of the file
path = "gfg.xlsx"

# To open the workbook
# workbook object is created
wb_obj = openpyxl.load_workbook(path)

# Get workbook active sheet object
# from the active attribute
sheet_obj = wb_obj.active

# Cell objects also have a row, column,
# and coordinate attributes that provide
# location information for the cell.

# Note: The first row or
# column integer is 1, not 0.

# Cell object is created by using
# sheet object's cell() method.
cell_obj = sheet_obj.cell(row = 1, column = 1)

# Print value of cell object
# using the value attribute
print(cell_obj.value)

from openpyxl import Workbook

# Call a Workbook() function of openpyxl
# to create a new blank Workbook object
workbook = Workbook()

# Anytime you modify the Workbook object
# or its sheets and cells, the spreadsheet
# file will not be saved until you call
# the save() workbook method.
workbook.save(filename="sample.xlsx")

# import openpyxl module
import openpyxl

# Call a Workbook() function of openpyxl
# to create a new blank Workbook object
wb = openpyxl.Workbook()

# Get workbook active sheet
# from the active attribute
sheet = wb.active

# Cell objects also have row, column
# and coordinate attributes that provide
# location information for the cell.

# Note: The first row or column integer
# is 1, not 0. Cell object is created by
# using sheet object's cell() method.
c1 = sheet.cell(row = 1, column = 1)

# writing values to cells
c1.value = "Hello"

c2 = sheet.cell(row= 1 , column = 2)
c2.value = "World"

# Once have a Worksheet object, one can
# access a cell object by its name also.
# A2 means column = 1 & row = 2.
c3 = sheet['A2']
c3.value = "Welcome"

# B2 means column = 2 & row = 2.
c4 = sheet['B2']
c4.value = "Everyone"

# Anytime you modify the Workbook object
# or its sheets and cells, the spreadsheet
# file will not be saved until you call
# the save() workbook method.
wb.save("sample.xlsx")

# import openpyxl module
import openpyxl
