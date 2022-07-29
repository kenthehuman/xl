from openpyxl import Workbook

workbook = Workbook()
sheet = workbook.active

sheet["A1"] = "hello"
sheet["B1"] = "world!"

workbook.save(filename="hello_world.xlsx")



from openpyxl import load_workbook
workbook = load_workbook(filename="sample.xlsx")
workbook.sheetnames


sheet = workbook.active
sheet


sheet.title

>>> sheet.cell(row=10, column=6)
<Cell 'Sheet 1'.F10>

>>> sheet.cell(row=10, column=6).value
"G-Shock Men's Grey Sport Watch"

>>> sheet["A1:C2"]
((<Cell 'Sheet 1'.A1>, <Cell 'Sheet 1'.B1>, <Cell 'Sheet 1'.C1>),
 (<Cell 'Sheet 1'.A2>, <Cell 'Sheet 1'.B2>, <Cell 'Sheet 1'.C2>))

>>> for value in sheet.iter_rows(min_row=1,
...                              max_row=2,
...                              min_col=1,
...                              max_col=3,
...                              values_only=True):
...     print(value)
('marketplace', 'customer_id', 'review_id')
('US', 3653882, 'R3O9SGZBVQBV76')

import json
from openpyxl import load_workbook

workbook = load_workbook(filename="sample.xlsx")
sheet = workbook.active

products = {}

# Using the values_only because you want to return the cells' values
for row in sheet.iter_rows(min_row=2,
                           min_col=4,
                           max_col=7,
                           values_only=True):
    product_id = row[0]
    product = {
        "parent": row[1],
        "title": row[2],
        "category": row[3]
    }
    products[product_id] = product

# Using json here to be able to format the output for displaying later
print(json.dumps(products))


import json
from openpyxl import load_workbook

workbook = load_workbook(filename="sample.xlsx")
sheet = workbook.active

products = {}

# Using the values_only because you want to return the cells' values
for row in sheet.iter_rows(min_row=2,
                           max_row=4,
                           min_col=4,
                           max_col=7,
                           values_only=True):
    product_id = row[0]
    product = {
        "parent": row[1],
        "title": row[2],
        "category": row[3]
    }
    products[product_id] = product

# Using json here to be able to format the output for displaying later
print(json.dumps(products))


# 
# 
# Convert Data Into Python Classes

# classes.py
import datetime
from dataclasses import dataclass

@dataclass
class Product:
    id: str
    parent: str
    title: str
    category: str

@dataclass
class Review:
    id: str
    customer_id: str
    stars: int
    headline: str
    body: str
    date: datetime.datetime



for value in sheet.iter_rows(min_row=1,
                             max_row=1,
                             values_only=True):
    print(value)


# Or an alternative
for cell in sheet[1]:
    print(cell.value)


>>> for value in sheet.iter_rows(min_row=1,
...                              max_row=1,
...                              values_only=True):
...     print(value)
('marketplace', 'customer_id', 'review_id', 'product_id', ...)



# Or an alternative
for cell in sheet[1]:
    print(cell.value)
marketplace
customer_id
review_id
product_id
product_parent
...


# mapping.py
# Product fields
PRODUCT_ID = 3
PRODUCT_PARENT = 4
PRODUCT_TITLE = 5
PRODUCT_CATEGORY = 6

# Review fields
REVIEW_ID = 2
REVIEW_CUSTOMER = 1
REVIEW_STARS = 7
REVIEW_HEADLINE = 12
REVIEW_BODY = 13
REVIEW_DATE = 14



from datetime import datetime
from openpyxl import load_workbook
from classes import Product, Review
from mapping import PRODUCT_ID, PRODUCT_PARENT, PRODUCT_TITLE, \
    PRODUCT_CATEGORY, REVIEW_DATE, REVIEW_ID, REVIEW_CUSTOMER, \
    REVIEW_STARS, REVIEW_HEADLINE, REVIEW_BODY

# Using the read_only method since you're not gonna be editing the spreadsheet
workbook = load_workbook(filename="sample.xlsx", read_only=True)
sheet = workbook.active

products = []
reviews = []

# Using the values_only because you just want to return the cell value
for row in sheet.iter_rows(min_row=2, values_only=True):
    product = Product(id=row[PRODUCT_ID],
                        parent=row[PRODUCT_PARENT],
                        title=row[PRODUCT_TITLE],
                        category=row[PRODUCT_CATEGORY])
    products.append(product)
    # You need to parse the date from the spreadsheet into a datetime format
    spread_date = row[REVIEW_DATE]
    # parsed_date = datetime.strftispreadme(spread_date, "%Y-%m-%d")
    review = Review(id=row[REVIEW_ID],
                    customer_id=row[REVIEW_CUSTOMER],
                    stars=row[REVIEW_STARS],
                    headline=row[REVIEW_HEADLINE],
                    body=row[REVIEW_BODY],
                    date=spread_date.strftime("%Y-%m-%d"))
    reviews.append(review)
print(product[0])
print(reviews[0])

# date=datetime.datetime(2015, 8, 31, 0, 0)) not sure why date = this . this also did same parsed_date = datetime.strptime(spread_date, "%Y-%m-%d")



# Appending New Data
from openpyxl import load_workbook

# Start by opening the spreadsheet and selecting the main sheet
workbook = load_workbook(filename="hello_world.xlsx")
sheet = workbook.active

# Write what you want into a specific cell
sheet["C1"] = "writing ;)"

# Save the spreadsheet
workbook.save(filename="hello_world_append.xlsx")


# Creating a Simple Spreadsheet
from openpyxl import Workbook

filename = "hello_world.xlsx"

workbook = Workbook()
sheet = workbook.active

sheet["A1"] = "hello"
sheet["B1"] = "world!"

workbook.save(filename=filename)

# print each row
def print_rows():
    for row in sheet.iter_rows(values_only=True):
        print(row)